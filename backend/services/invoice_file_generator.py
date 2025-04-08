import io
import re
import os
import tempfile
import pythoncom
import win32com.client
from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from sqlalchemy.orm import Session
from backend.database import MainSessionLocal
from backend.models import Project, Subtask, Invoice, InvoiceAmount



def generate_invoice_excel_bytes(project_id: int, invoice_number: str, template_path: str) -> bytes:
    db: Session = MainSessionLocal()
    try:
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            raise Exception("Project not found")

        invoice = db.query(Invoice).filter(Invoice.invoice_number == invoice_number).first()
        if not invoice:
            raise Exception("Invoice not found")

        invoice_items = db.query(InvoiceAmount).filter(InvoiceAmount.invoice_id == invoice.id).all()

        subtask_ids = [item.subtask_id for item in invoice_items]
        subtasks = db.query(Subtask).filter(Subtask.id.in_(subtask_ids)).all()

        # Build a lookup
        subtask_lookup = {sub.id: sub for sub in subtasks}

        match = re.match(r"(.+)-(\d+)", invoice_number)
        if match:
            base_number, current_suffix = match.groups()
            prev_suffix = int(current_suffix) - 1
            previous_invoice_number = f"{base_number}-{prev_suffix}" if prev_suffix > 0 else None
        else:
            previous_invoice_number = None

        # Get previous invoice total
        prev_invoice_total = 0.00
        if previous_invoice_number:
            prev_invoices = db.query(Invoice).filter(Invoice.invoice_number == previous_invoice_number).all()
            prev_invoice_total = sum(
                sum(Decimal(str(item.invoice_amount)) for item in inv.invoice_items)
                for inv in prev_invoices
            ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        wb = load_workbook(template_path)
        ws = wb.active

        # Header
        ws["K48"] = float(prev_invoice_total)
        ws["C3"] = project.bmcd_number
        ws["C4"] = invoice_number
        ws["C5"] = project.contract_number
        ws["C6"] = project.po_number
        ws["C7"] = f"{project.tao_number}-{project.po_number}"
        ws["C8"] = project.wo_number
        ws["C9"] = project.project_name
        ws["C10"] = float(project.total_budget_amount or 0)

        def generate_cell_map():
            base_row_map = {
                "labor": 15,
                "fee": 16,
                "non-travel expenses": 17,
                "travel expenses": 18,
            }
            cell_map = {}

            for subtask_name, row_offset in [("Scoping", 0), ("Physical", 7), ("P&C", 14), ("Telecom", 21)]:
                for category, base_row in base_row_map.items():
                    row = base_row + row_offset
                    cell_map[(subtask_name, category)] = (f"H{row}", f"K{row}")

            return cell_map


        formatted_date = datetime.strptime(invoice.invoice_through_date, "%Y-%m-%d").strftime("%d-%b-%y")
        ws["J3"] = formatted_date

        cell_map = generate_cell_map()

        for item in invoice_items:
            sub = subtask_lookup[item.subtask_id]
            key = (sub.subtask_name, sub.budget_category.strip().lower())
            if key in cell_map:
                cell_line_item, cell_amount = cell_map[key]
                ws[cell_line_item] = sub.line_item
                ws[cell_amount] = float(item.invoice_amount or 0)

        # Save to memory
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        return file_stream.read()

    finally:
        db.close()

def generate_invoice_pdf_bytes(project_id: int, invoice_number: str, template_path: str) -> bytes:
    # Step 1: Generate Excel bytes in memory
    excel_bytes = generate_invoice_excel_bytes(project_id, invoice_number, template_path)

    # Step 2: Write Excel to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
        tmp_excel.write(excel_bytes)
        tmp_excel_path = tmp_excel.name

    tmp_pdf_path = tmp_excel_path.replace(".xlsx", ".pdf")

    # Step 3: Use COM to convert to PDF
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False

        wb = excel.Workbooks.Open(tmp_excel_path)
        wb.ExportAsFixedFormat(0, tmp_pdf_path)  # 0 = PDF format
        wb.Close(False)
        excel.Quit()
    finally:
        pythoncom.CoUninitialize()

    try:
        with open(tmp_pdf_path, "rb") as f:
            pdf_bytes = f.read()
    except Exception as e:
        raise Exception(f"Failed to read generated PDF: {e}")

    try:
        os.remove(tmp_excel_path)
        os.remove(tmp_pdf_path)
    except Exception as cleanup_error:
        print(f"Warning: Failed to clean up temp files: {cleanup_error}")

    return pdf_bytes
